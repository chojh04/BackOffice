# -- coding:utf-8 --
'''
Created on 2017. 3. 13.

@author: sanghyun
'''
from collections import OrderedDict
import datetime
import json
from mhlib import isnumeric
import os
import sys
import threading
import time
import zipfile

from flask import Blueprint, request
from flask.globals import current_app, session
import openpyxl
import xlsxwriter

from routes.api.systemMngApi import postBatchMng, putBatchMng
from util.common import  paramEscape, getApiData, getParameter, \
    setStringToNumber, getData, EXCEL_FILE_DOWNLOAD_COUNT, parseDate, \
    EXCEL_FILE_MAKE_LIMT_COUNT, setUnicodeEncodeTypeToEucKr, setUnicodeFormatToEucKr



approvalsApi = Blueprint('approvalsApi', __name__)

@approvalsApi.route('/api/approvals/payments', methods=['GET'])
def payments():
    form_data = json.loads(request.args.get("formData"))
    searchDate = getParameter(form_data , "startDate").split(' - ')
    selType = getParameter(form_data , "selType")
    startDate = paramEscape(searchDate[0])
    endDate = paramEscape(searchDate[1])    
    serviceNm = ""
    merchantNm = ""
    if selType == "1" :
        serviceNm = getParameter(form_data , "selName")
    else :
        merchantNm = getParameter(form_data , "selName")
    queryData = {
        'summSeq':'',
        'merchantId': getParameter(form_data , "submerchantId"),
        'serviceId': getParameter(form_data , "serviceId"),
        'approvalStatus': getParameter(form_data , "status"),
        'orderNo': getParameter(form_data,"orderNo"),
        'approvalNo' :getParameter(form_data,"approvalNo"),
        'dateType': getParameter(form_data,"dateType"),
        'svcConnId': getParameter(form_data,"svcConnId"),
        'startDate': startDate,
        'endDate': endDate,
        'merchantNm': merchantNm,
        'serviceNm': serviceNm,
        'offset': setStringToNumber(request.args.get("start")),
        'limit': setStringToNumber(request.args.get("length")),
        'sort':'',
        'clause':'',
        'payMethod': getParameter(form_data,"payMethod"),
        'saleDivider': getParameter(form_data,"saleDivider"),
        'cardType': getParameter(form_data,"cardType"),
        'excelAllFlag':'',
    }
    result_data = getApiData("/approvals/payments" ,queryData)
    return json.dumps(result_data
                      )
@approvalsApi.route('/api/approvals/paymentsDaily', methods=['GET'])
def paymentsDaily():
    form_data = json.loads(request.args.get("formData"))
    searchDate = getParameter(form_data , "startDate").split(' - ')
    selType = getParameter(form_data , "selType")
    startDate = paramEscape(searchDate[0])
    endDate = paramEscape(searchDate[1])    
    serviceNm = ""
    merchantNm = ""
    if selType == "1" :
        serviceNm = getParameter(form_data , "selName")
    else :
        merchantNm = getParameter(form_data , "selName")
    queryData = {
        'summSeq':'',
        'merchantId': getParameter(form_data , "submerchantId"),
        'serviceId': getParameter(form_data , "serviceId"),
        'approvalStatus': getParameter(form_data , "status"),
        'orderNo': getParameter(form_data,"orderNo"),
        'approvalNo' :getParameter(form_data,"approvalNo"),
        'dateType': getParameter(form_data,"dateType"),
        'svcConnId': getParameter(form_data,"svcConnId"),
        'startDate': startDate,
        'endDate': endDate,
        'merchantNm': merchantNm,
        'serviceNm': serviceNm,
        'offset': setStringToNumber(request.args.get("start")),
        'limit': setStringToNumber(request.args.get("length")),
        'sort':'',
        'clause':'',
        'payMethod': getParameter(form_data,"payMethod"),
        'saleDivider': getParameter(form_data,"saleDivider"),
        'cardType': getParameter(form_data,"cardType"),
        'excelAllFlag':'',
    }
    result_data = getApiData("/approvals/paymentsDaily" ,queryData)
    return json.dumps(result_data)

@approvalsApi.route('/api/approvals/paymentsMonth', methods=['GET'])
def paymentsMonth():
    form_data = json.loads(request.args.get("formData"))
    searchDate = getParameter(form_data , "startDate").split(' - ')
    selType = getParameter(form_data , "selType")
    startDate = paramEscape(searchDate[0])
    endDate = paramEscape(searchDate[1])    
    serviceNm = ""
    merchantNm = ""
    if selType == "1" :
        serviceNm = getParameter(form_data , "selName")
    else :
        merchantNm = getParameter(form_data , "selName")
    queryData = {
        'summSeq':'',
        'merchantId': getParameter(form_data , "submerchantId"),
        'serviceId': getParameter(form_data , "serviceId"),
        'approvalStatus': getParameter(form_data , "status"),
        'orderNo': getParameter(form_data,"orderNo"),
        'approvalNo' :getParameter(form_data,"approvalNo"),
        'dateType': getParameter(form_data,"dateType"),
        'svcConnId': getParameter(form_data,"svcConnId"),
        'startDate': startDate,
        'endDate': endDate,
        'merchantNm': merchantNm,
        'serviceNm': serviceNm,
        'offset': setStringToNumber(request.args.get("start")),
        'limit': setStringToNumber(request.args.get("length")),
        'sort':'',
        'clause':'',
        'payMethod': getParameter(form_data,"payMethod"),
        'saleDivider': getParameter(form_data,"saleDivider"),
        'cardType': getParameter(form_data,"cardType"),
        'excelAllFlag':'',
    }
    result_data = getApiData("/approvals/paymentsMonth" ,queryData)
    return json.dumps(result_data)

@approvalsApi.route('/api/approvals/payments/excelAll', methods=['GET'])
def paymentsExcelAll():
    searchDate = getParameter({} , "startDate").split(' - ')
    selType = getParameter({}, "selType")
    startDate = paramEscape(searchDate[0])
    endDate = paramEscape(searchDate[1])    
    serviceNm = ""
    merchantNm = ""
    if selType == "1" :
        serviceNm = getParameter({} , "selName")
    else :
        merchantNm = getParameter({} , "selName")
    queryData = {
        'summSeq':'',
        'merchantId': getParameter({} , "submerchantId"),
        'serviceId': getParameter({} , "serviceId"),
        'approvalStatus': getParameter({} , "status"),
        'orderNo': getParameter({},"orderNo"),
        'approvalNo' :getParameter({},"approvalNo"),
        'dateType': getParameter({},"dateType"),
        'svcConnId': getParameter({},"svcConnId"),
        'startDate': startDate,
        'endDate': endDate,
        'merchantNm': merchantNm,
        'serviceNm': serviceNm,
        'offset': 0,
        'limit': EXCEL_FILE_DOWNLOAD_COUNT,
        'sort':'',
        'clause':'',
        'excelAllFlag':'1',
        'payMethod': getParameter({},"payMethod"),
        'saleDivider': getParameter({},"saleDivider"),
        'cardType': getParameter({},"cardType"),        
        'empId' : session['empId']
    }
    rootPath = current_app.root_path
    t1 = threading.Thread(target=makePaymentsExcelFile,args=[queryData,rootPath])
    t1.daemon = True
    t1.start()
    
    return "엑셀 작업요청"

def makePaymentsExcelFile(queryData,rootPath):
    excelZip = None
    jobStatus = 0
    batchId = None
    try:
        fileCnt = 1
        makeTime = str(int(round(time.time()*1000)))
        uploads = os.path.join(rootPath, "fileDownload" , "excel" , makeTime)
        if not os.path.isdir(uploads):
            os.makedirs(uploads)
        zipFileName = u'거래내역_결제_'+ datetime.datetime.now().strftime('%Y%m%d') +'.zip'
        #작업 시작 메시지 추가
        batchId = postBatchMng({
            "reqId"    : queryData['empId'],
            "status"   : "BAT-0001" , # 진행중
            "filePath" : os.path.join(uploads ,zipFileName),
            "content"  : "결제 정산 엑셀 배치작업",
            "errMsg"   : ""
        })["data"]["batchId"]            
        fileName = '거래내역_결제_' + datetime.datetime.now().strftime('%Y%m%d') + '_' +  str(fileCnt) +  '.xlsx'
        workbook = xlsxwriter.Workbook(os.path.join(uploads ,setUnicodeEncodeTypeToEucKr(fileName)))
        worksheet = workbook.add_worksheet()
        
        money_format = workbook.add_format()        
        money_format.set_num_format('_- #,##0_-;[red]- #,##0_-;_- "-"_-;_-@_-')
        ratio_format = workbook.add_format()
        ratio_format.set_num_format('_- #,##0.0"%"_-;[red]- #,##0.0"%"_-;_- "-"_-;_-@_-')
        
        row = 0  
        worksheet.write(row, 0  ,"거래처명")
        worksheet.write(row, 1  ,"서비스명")
        worksheet.write(row, 2  ,"연동아이디")
        worksheet.write(row, 3  ,"점포코드")
        worksheet.write(row, 4  ,"점포명(사용처)")
        worksheet.write(row, 5  ,"상태")
        worksheet.write(row, 6  ,"지불수단")
        worksheet.write(row, 7  ,"지불형태")
        worksheet.write(row, 8  ,"카드구분")
        worksheet.write(row, 9  ,"영업일")
        worksheet.write(row, 10  ,"거래일")
        worksheet.write(row, 11 ,"거래시간")
        worksheet.write(row, 12 ,"주문번호")
        worksheet.write(row, 13 ,"승인번호")
        worksheet.write(row, 14 ,"카드번호")
        worksheet.write(row, 15 ,"거래금액")
        worksheet.write(row, 16 ,"결제금액")
        worksheet.write(row, 17 ,"할인금액")
        worksheet.write(row, 18 ,"수수료")
        worksheet.write(row, 19 ,"결제처 수수료율")
        worksheet.write(row, 20 ,"결제처 수수료")
        worksheet.write(row, 21 ,"고객지급금액")
        worksheet.write(row, 22 ,"비고")
        
        while True : 
            searchData = getData("/approvals/payments" ,queryData)
            for data in searchData["resultList"]:
                row += 1
                amount = long(data["amount"])
                dcAmount = long(data["dcAmount"])
                origAmount = long(data["origAmount"])
                commision = 0
                kpcCommision = 0
          
                if data["commision"] > 0 and isnumeric(data["commision"]):
                    commision = int(data["commision"])
                if data["kpcCommision"] > 0 :
                    kpcCommision = int(data["kpcCommision"]);
                if data["approvalStatus"] == "PAYT-0002" or data["approvalStatus"] == "PAYT-0004":
                    amount *= -1
                    dcAmount *= -1
                    origAmount *= -1
                    if commision > 0 :
                        commision *= -1
                    if kpcCommision > 0 :
                        kpcCommision *= -1
                        
                worksheet.write(row, 0  ,data["submerchantNm"])
                worksheet.write(row, 1  ,data["serviceNm"])
                worksheet.write(row, 2  ,data["svcConnId"])
                worksheet.write(row, 3  ,data["storeCd"])
                worksheet.write(row, 4  ,data["storeNm"])
                worksheet.write(row, 5  ,data["approvalStatusNm"])
                worksheet.write(row, 6  ,data["payMethodName"])
                worksheet.write(row, 7  ,data["cardCdName"])
                worksheet.write(row, 8  ,data["saleDividerName"])
                worksheet.write(row, 9  ,parseDate(data["dealDt"] ,'%Y%m%d','%Y-%m-%d'))
                worksheet.write(row, 10 ,parseDate(data["approvalDt"] ,'%Y-%m-%d %H:%M:%S' ,'%Y-%m-%d'))
                worksheet.write(row, 11 ,parseDate(data["approvalDt"] ,'%Y-%m-%d %H:%M:%S' ,'%H:%M:%S'))
                worksheet.write(row, 12 ,data["orderNo"])
                worksheet.write(row, 13 ,data["approvalNo"])
                worksheet.write(row, 14 ,data["cardEncNo"])
                worksheet.write_number(row, 15, origAmount, money_format)
                worksheet.write_number(row, 16, amount, money_format)
                worksheet.write_number(row, 17, dcAmount, money_format)
                worksheet.write_number(row, 18, commision, money_format)
                worksheet.write_number(row, 19, data["kpcCommisionRatio"], ratio_format)
                worksheet.write_number(row, 20, kpcCommision, money_format)
                worksheet.write_number(row, 21, data["customerProvideAmount"], money_format)
                worksheet.write(row, 22,data["desc01"])
                if row >= EXCEL_FILE_MAKE_LIMT_COUNT :
                    row = 0
                    fileCnt += 1
                    fileName = '거래내역_결제_' +  datetime.datetime.now().strftime('%Y%m%d') + '_' + str(fileCnt) +  '.xlsx'
                    # 디비 조회건수 * 2 row 생성시 파일 재생성
                    workbook.close()
                    workbook = xlsxwriter.Workbook(os.path.join(uploads ,setUnicodeEncodeTypeToEucKr(fileName)))
                    worksheet = workbook.add_worksheet()
                    money_format = workbook.add_format()
                    money_format.set_num_format('_- #,##0_-;[red]- #,##0_-;_- "-"_-;_-@_-')
                    worksheet.write(row, 0  ,"거래처명")
                    worksheet.write(row, 1  ,"서비스명")
                    worksheet.write(row, 2  ,"연동아이디")
                    worksheet.write(row, 3  ,"점포코드")
                    worksheet.write(row, 4  ,"점포명(사용처)")
                    worksheet.write(row, 5  ,"상태")
                    worksheet.write(row, 6  ,"지불수단")
                    worksheet.write(row, 7  ,"지불형태")
                    worksheet.write(row, 8  ,"카드구분")
                    worksheet.write(row, 9  ,"영업일")
                    worksheet.write(row, 10  ,"거래일")
                    worksheet.write(row, 11 ,"거래시간")
                    worksheet.write(row, 12 ,"주문번호")
                    worksheet.write(row, 13 ,"승인번호")
                    worksheet.write(row, 14 ,"카드번호")
                    worksheet.write(row, 15 ,"거래금액")
                    worksheet.write(row, 16 ,"결제금액")
                    worksheet.write(row, 17 ,"할인금액")
                    worksheet.write(row, 18 ,"수수료")
                    worksheet.write(row, 19 ,"결제처 수수료율")
                    worksheet.write(row, 20 ,"결제처 수수료")
                    worksheet.write(row, 21 ,"고객지급금액")
                    worksheet.write(row, 22 ,"비고")                           
            queryData["offset"] = queryData["offset"] + EXCEL_FILE_DOWNLOAD_COUNT 
            if len(searchData["resultList"]) < EXCEL_FILE_DOWNLOAD_COUNT : 
                break
        workbook.close()
        excelZip = zipfile.ZipFile(os.path.join(uploads ,zipFileName),'w')
        for folder, subfolders, files in os.walk(uploads):
            for file in files:
                if file.endswith('.xlsx'):
                    excelZip.write(os.path.join(folder ,file), setUnicodeFormatToEucKr(file), compress_type=zipfile.ZIP_DEFLATED)
                    
    except Exception as err:
        #에러 메시지 추가
        putBatchMng({
            "batchId"  : str(batchId),
            "reqId"    : queryData['empId'],
            "status"   : "BAT-0003" , # 진행중
            "errMsg"   : str(err)
        })
        jobStatus = 1
        print err
    finally:
        if excelZip != None:
            excelZip.close()
        if jobStatus == 0 :    
            putBatchMng({
                "batchId"  : str(batchId),
                "reqId"    : queryData['empId'],
                "status"   : "BAT-0002" , # 진행중
                "errMsg"   : ""
            })               
        #성공 메시지 추가
        print "성공"

@approvalsApi.route('/api/approvals/payments/dailyExcelAll', methods=['GET'])
def paymentsDailyExcelAll():
    searchDate = getParameter({} , "startDate").split(' - ')
    selType = getParameter({}, "selType")
    startDate = paramEscape(searchDate[0])
    endDate = paramEscape(searchDate[1])    
    serviceNm = ""
    merchantNm = ""
    if selType == "1" :
        serviceNm = getParameter({} , "selName")
    else :
        merchantNm = getParameter({} , "selName")
    queryData = {
        'summSeq':'',
        'merchantId': getParameter({} , "submerchantId"),
        'serviceId': getParameter({} , "serviceId"),
        'approvalStatus': getParameter({} , "status"),
        'orderNo': getParameter({},"orderNo"),
        'approvalNo' :getParameter({},"approvalNo"),
        'dateType': getParameter({},"dateType"),
        'svcConnId': getParameter({},"svcConnId"),
        'startDate': startDate,
        'endDate': endDate,
        'merchantNm': merchantNm,
        'serviceNm': serviceNm,
        'offset': 0,
        'limit': EXCEL_FILE_DOWNLOAD_COUNT,
        'sort':'',
        'clause':'',
        'excelAllFlag':'1',
        'payMethod': getParameter({},"payMethod"),
        'saleDivider': getParameter({},"saleDivider"),
        'cardType': getParameter({},"cardType"),        
        'empId' : session['empId']
    }
    rootPath = current_app.root_path
    t1 = threading.Thread(target=makePaymentsDailyExcelFile,args=[queryData,rootPath])
    t1.daemon = True
    t1.start()
    return "엑셀 작업요청"

def makePaymentsDailyExcelFile(queryData,rootPath):
    excelZip = None
    jobStatus = 0
    batchId = None
    try:
        fileCnt = 1
        makeTime = str(int(round(time.time()*1000)))
        uploads = os.path.join(rootPath, "fileDownload" , "excel" , makeTime)
        if not os.path.isdir(uploads):
            os.makedirs(uploads)
        zipFileName = u'거래내역_결제_일별_'+ datetime.datetime.now().strftime('%Y%m%d') +'.zip'
        #작업 시작 메시지 추가
        batchId = postBatchMng({
            "reqId"    : queryData['empId'],
            "status"   : "BAT-0001" , # 진행중
            "filePath" : os.path.join(uploads ,zipFileName),
            "content"  : "결제 일별 정산 엑셀 배치작업",
            "errMsg"   : ""
        })["data"]["batchId"]            
        fileName = '거래내역_결제_일별_' +  datetime.datetime.now().strftime('%Y%m%d') + '_' + str(fileCnt) +  '.xlsx'
        workbook = xlsxwriter.Workbook(os.path.join(uploads ,setUnicodeEncodeTypeToEucKr(fileName)))
        worksheet = workbook.add_worksheet()
        money_format = workbook.add_format()
        money_format.set_num_format('_- #,##0_-;[red]- #,##0_-;_- "-"_-;_-@_-')
        row = 0
        worksheet.write(row, 0  ,"거래처명")
        worksheet.write(row, 1  ,"서비스명")
        worksheet.write(row, 2  ,"연동아이디")
        worksheet.write(row, 3  ,"상태")
        worksheet.write(row, 4  ,"지불수단")
        worksheet.write(row, 5  ,"카드구분")
        worksheet.write(row, 6  ,"영업일")
        worksheet.write(row, 7  ,"거래일")
        worksheet.write(row, 8  ,"거래건수")
        worksheet.write(row, 9  ,"거래금액")
        worksheet.write(row, 10 ,"수수료")
        while True : 
            searchData = getData("/approvals/paymentsDaily" ,queryData)
            for data in searchData["resultList"]:
                row += 1
                amount = long(data["amount"])
                commision = 0
                if data["commision"] > 0 and isnumeric(data["commision"]):
                    commision = int(data["commision"])
                if data["approvalStatus"] == "PAYT-0002" or data["approvalStatus"] == "PAYT-0004":
                    amount *= -1
                    if commision > 0 :
                        commision *= -1
                worksheet.write(row, 0  ,data["submerchantNm"])
                worksheet.write(row, 1  ,data["serviceNm"])
                worksheet.write(row, 2  ,data["svcConnId"])
                worksheet.write(row, 3  ,data["approvalStatusNm"])
                worksheet.write(row, 4  ,data["payTypeName"])
                worksheet.write(row, 5  ,data["cardCdName"])
                worksheet.write(row, 6  ,parseDate(data["dealDt"] ,'%Y%m%d','%Y-%m-%d'))
                worksheet.write(row, 7  ,parseDate(data["dealDt"] ,'%Y%m%d','%Y-%m-%d'))
                worksheet.write_number(row, 8, long(data["cnt"]), money_format)
                worksheet.write_number(row, 9, amount, money_format)
                worksheet.write_number(row, 10, long(commision), money_format)
                if row >= EXCEL_FILE_MAKE_LIMT_COUNT :
                    row = 0
                    fileCnt += 1
                    fileName = '거래내역_결제_일별_' +  datetime.datetime.now().strftime('%Y%m%d') + '_'  +  str(fileCnt) +  '.xlsx'
                    # 디비 조회건수 * 2 row 생성시 파일 재생성
                    workbook.close()
                    workbook = xlsxwriter.Workbook(os.path.join(uploads ,setUnicodeEncodeTypeToEucKr(fileName)))
                    worksheet = workbook.add_worksheet()
                    worksheet.write(row, 0  ,"거래처명")
                    worksheet.write(row, 1  ,"서비스명")
                    worksheet.write(row, 2  ,"연동아이디")
                    worksheet.write(row, 3  ,"상태")
                    worksheet.write(row, 4  ,"지불수단")
                    worksheet.write(row, 5  ,"카드구분")
                    worksheet.write(row, 6  ,"영업일")
                    worksheet.write(row, 7  ,"거래일")
                    worksheet.write(row, 8  ,"거래건수")
                    worksheet.write(row, 9  ,"거래금액")
                    worksheet.write(row, 10 ,"수수료")               
            queryData["offset"] = queryData["offset"] + EXCEL_FILE_DOWNLOAD_COUNT 
            if len(searchData["resultList"]) < EXCEL_FILE_DOWNLOAD_COUNT : 
                break
        workbook.close()
        excelZip = zipfile.ZipFile(os.path.join(uploads ,zipFileName),'w')
        for folder, subfolders, files in os.walk(uploads):
            for file in files:
                if file.endswith('.xlsx'):
                    excelZip.write(os.path.join(folder ,file), setUnicodeFormatToEucKr(file), compress_type=zipfile.ZIP_DEFLATED)
                    
    except Exception as err:
        #에러 메시지 추가
        putBatchMng({
            "batchId"  : str(batchId),
            "reqId"    : queryData['empId'],
            "status"   : "BAT-0003" , # 진행중
            "errMsg"   : str(err)
        })
        jobStatus = 1
        print err
    finally:
        if excelZip != None:
            excelZip.close()
        if jobStatus == 0 :    
            putBatchMng({
                "batchId"  : str(batchId),
                "reqId"    : queryData['empId'],
                "status"   : "BAT-0002" , # 진행중
                "errMsg"   : ""
            })               
        #성공 메시지 추가
        print "성공"
        
@approvalsApi.route('/api/approvals/payments/monthExcelAll', methods=['GET'])
def paymentsMonthExcelAll():
    searchDate = getParameter({} , "startDate").split(' - ')
    selType = getParameter({}, "selType")
    startDate = paramEscape(searchDate[0])
    endDate = paramEscape(searchDate[1])    
    serviceNm = ""
    merchantNm = ""
    if selType == "1" :
        serviceNm = getParameter({} , "selName")
    else :
        merchantNm = getParameter({} , "selName")
    queryData = {
        'summSeq':'',
        'merchantId': getParameter({} , "submerchantId"),
        'serviceId': getParameter({} , "serviceId"),
        'approvalStatus': getParameter({} , "status"),
        'orderNo': getParameter({},"orderNo"),
        'approvalNo' :getParameter({},"approvalNo"),
        'dateType': getParameter({},"dateType"),
        'svcConnId': getParameter({},"svcConnId"),
        'startDate': startDate,
        'endDate': endDate,
        'merchantNm': merchantNm,
        'serviceNm': serviceNm,
        'offset': 0,
        'limit': EXCEL_FILE_DOWNLOAD_COUNT,
        'sort':'',
        'clause':'',
        'excelAllFlag':'1',
        'payMethod': getParameter({},"payMethod"),
        'saleDivider': getParameter({},"saleDivider"),
        'cardType': getParameter({},"cardType"),        
        'empId' : session['empId']
    }
    rootPath = current_app.root_path
    t1 = threading.Thread(target=makePaymentsMonthExcelFile,args=[queryData,rootPath])
    t1.daemon = True
    t1.start()
    return "엑셀 작업요청"

def makePaymentsMonthExcelFile(queryData,rootPath):
    excelZip = None
    jobStatus = 0
    batchId = None
    try:
        fileCnt = 1
        makeTime = str(int(round(time.time()*1000)))
        uploads = os.path.join(rootPath, "fileDownload" , "excel" , makeTime)
        if not os.path.isdir(uploads):
            os.makedirs(uploads)
        zipFileName = u'거래내역_결제_월별_'+ datetime.datetime.now().strftime('%Y%m%d') +'.zip'
        #작업 시작 메시지 추가
        batchId = postBatchMng({
            "reqId"    : queryData['empId'],
            "status"   : "BAT-0001" , # 진행중
            "filePath" : os.path.join(uploads ,zipFileName),
            "content"  : "결제 월별 정산 엑셀 배치작업",
            "errMsg"   : ""
        })["data"]["batchId"]            
        fileName = '거래내역_결제_월별_' +  datetime.datetime.now().strftime('%Y%m%d') + '_' + str(fileCnt) +  '.xlsx'
        workbook = xlsxwriter.Workbook(os.path.join(uploads ,setUnicodeEncodeTypeToEucKr(fileName)))
        worksheet = workbook.add_worksheet()
        money_format = workbook.add_format()
        money_format.set_num_format('_- #,##0_-;[red]- #,##0_-;_- "-"_-;_-@_-')
        row = 0
        worksheet.write(row, 0  ,"거래처명")
        worksheet.write(row, 1  ,"서비스명")
        worksheet.write(row, 2  ,"연동아이디")
        worksheet.write(row, 3  ,"상태")
        worksheet.write(row, 4  ,"지불수단")
        worksheet.write(row, 5  ,"카드구분")
        worksheet.write(row, 6  ,"영업일")
        worksheet.write(row, 7  ,"거래일")
        worksheet.write(row, 8  ,"거래건수")
        worksheet.write(row, 9  ,"거래금액")
        worksheet.write(row, 10 ,"수수료")
        while True : 
            searchData = getData("/approvals/paymentsMonth" ,queryData)
            for data in searchData["resultList"]:
                row += 1
                amount = long(data["amount"])
                commision = 0
                if data["commision"] > 0 and isnumeric(data["commision"]):
                    commision = int(data["commision"])
                if data["approvalStatus"] == "PAYT-0002" or data["approvalStatus"] == "PAYT-0004":
                    amount *= -1
                    if commision > 0 :
                        commision *= -1
                worksheet.write(row, 0  ,data["submerchantNm"])
                worksheet.write(row, 1  ,data["serviceNm"])
                worksheet.write(row, 2  ,data["svcConnId"])
                worksheet.write(row, 3  ,data["approvalStatusNm"])
                worksheet.write(row, 4  ,data["payTypeName"])
                worksheet.write(row, 5  ,data["cardCdName"])
                worksheet.write(row, 6  ,parseDate(data["dealDt"] ,'%Y%m','%Y-%m'))
                worksheet.write(row, 7  ,parseDate(data["dealDt"] ,'%Y%m','%Y-%m'))
                worksheet.write_number(row, 8, long(data["cnt"]), money_format)
                worksheet.write_number(row, 9, amount, money_format)
                worksheet.write_number(row, 10, long(commision), money_format)
                if row >= EXCEL_FILE_MAKE_LIMT_COUNT :
                    row = 0
                    fileCnt += 1
                    fileName = '거래내역_결제_월별_' +  datetime.datetime.now().strftime('%Y%m%d') + '_' + str(fileCnt) +  '.xlsx'
                    # 디비 조회건수 * 2 row 생성시 파일 재생성
                    workbook.close()
                    workbook = xlsxwriter.Workbook(os.path.join(uploads ,setUnicodeEncodeTypeToEucKr(fileName)))
                    worksheet = workbook.add_worksheet()
                    worksheet.write(row, 0  ,"거래처명")
                    worksheet.write(row, 1  ,"서비스명")
                    worksheet.write(row, 2  ,"연동아이디")
                    worksheet.write(row, 3  ,"상태")
                    worksheet.write(row, 4  ,"지불수단")
                    worksheet.write(row, 5  ,"카드구분")
                    worksheet.write(row, 6  ,"영업일")
                    worksheet.write(row, 7  ,"거래일")
                    worksheet.write(row, 8  ,"거래건수")
                    worksheet.write(row, 9  ,"거래금액")
                    worksheet.write(row, 10 ,"수수료")                        
            queryData["offset"] = queryData["offset"] + EXCEL_FILE_DOWNLOAD_COUNT 
            if len(searchData["resultList"]) < EXCEL_FILE_DOWNLOAD_COUNT : 
                break
        workbook.close()
        excelZip = zipfile.ZipFile(os.path.join(uploads ,zipFileName),'w')
        for folder, subfolders, files in os.walk(uploads):
            for file in files:
                if file.endswith('.xlsx'):
                    excelZip.write(os.path.join(folder ,file), setUnicodeFormatToEucKr(file), compress_type=zipfile.ZIP_DEFLATED)
                    
    except Exception as err:
        #에러 메시지 추가
        putBatchMng({
            "batchId"  : str(batchId),
            "reqId"    : queryData['empId'],
            "status"   : "BAT-0003" , # 진행중
            "errMsg"   : str(err)
        })
        jobStatus = 1
        print err
    finally:
        if excelZip != None:
            excelZip.close()
        if jobStatus == 0 :    
            putBatchMng({
                "batchId"  : str(batchId),
                "reqId"    : queryData['empId'],
                "status"   : "BAT-0002" , # 진행중
                "errMsg"   : ""
            })               
        #성공 메시지 추가
        print "성공"        
    
@approvalsApi.route('/api/approvals/chargements', methods=['GET'])
def chargements():
    form_data = json.loads(request.args.get("formData"))
    searchDate = getParameter(form_data , "startDate").split(' - ')
    selType = getParameter(form_data , "selType")
    startDate = paramEscape(searchDate[0])
    endDate = paramEscape(searchDate[1])    
    serviceNm = ""
    merchantNm = ""
    if selType == "1" :
        serviceNm = getParameter(form_data , "selName")
    else :
        merchantNm = getParameter(form_data , "selName")
    queryData = {
        'summSeq':'',
        'merchantId': getParameter(form_data , "submerchantId"),
        'serviceId': getParameter(form_data , "serviceId"),
        'approvalStatus': getParameter(form_data , "status"),
        'orderNo': getParameter(form_data,"orderNo"),
        'approvalNo' :getParameter(form_data,"approvalNo"),
        'dateType': getParameter(form_data,"dateType"),
        'svcConnId': getParameter(form_data,"svcConnId"),
        'startDate': startDate,
        'endDate': endDate,
        'merchantNm': merchantNm,
        'serviceNm': serviceNm,
        'offset': setStringToNumber(request.args.get("start")),
        'limit': setStringToNumber(request.args.get("length")),
        'sort':'',
        'clause':'',
        'payMethod': getParameter(form_data,"payMethod"),
        'saleDivider': getParameter(form_data,"saleDivider"),
        'cardType': getParameter(form_data,"cardType"),        
        'excelAllFlag':'',
    }
#     print(queryData)
    result_data = getApiData("/approvals/chargements" ,queryData)
    return json.dumps(result_data)

@approvalsApi.route('/api/approvals/chargementsDaily', methods=['GET'])
def chargementsDaily():
    form_data = json.loads(request.args.get("formData"))
    searchDate = getParameter(form_data , "startDate").split(' - ')
    selType = getParameter(form_data , "selType")
    startDate = paramEscape(searchDate[0])
    endDate = paramEscape(searchDate[1])    
    serviceNm = ""
    merchantNm = ""
    if selType == "1" :
        serviceNm = getParameter(form_data , "selName")
    else :
        merchantNm = getParameter(form_data , "selName")
    queryData = {
        'summSeq':'',
        'merchantId': getParameter(form_data , "submerchantId"),
        'serviceId': getParameter(form_data , "serviceId"),
        'approvalStatus': getParameter(form_data , "status"),
        'orderNo': getParameter(form_data,"orderNo"),
        'approvalNo' :getParameter(form_data,"approvalNo"),
        'dateType': getParameter(form_data,"dateType"),
        'svcConnId': getParameter(form_data,"svcConnId"),
        'startDate': startDate,
        'endDate': endDate,
        'merchantNm': merchantNm,
        'serviceNm': serviceNm,
        'offset': setStringToNumber(request.args.get("start")),
        'limit': setStringToNumber(request.args.get("length")),
        'sort':'',
        'clause':'',
        'payMethod': getParameter(form_data,"payMethod"),
        'saleDivider': getParameter(form_data,"saleDivider"),
        'cardType': getParameter(form_data,"cardType"),        
        'excelAllFlag':'',
    }
    result_data = getApiData("/approvals/chargementsDaily" ,queryData)
    return json.dumps(result_data)

@approvalsApi.route('/api/approvals/chargementsMonth', methods=['GET'])
def chargementsMonth():
    form_data = json.loads(request.args.get("formData"))
    searchDate = getParameter(form_data , "startDate").split(' - ')
    selType = getParameter(form_data , "selType")
    startDate = paramEscape(searchDate[0])
    endDate = paramEscape(searchDate[1])    
    serviceNm = ""
    merchantNm = ""
    if selType == "1" :
        serviceNm = getParameter(form_data , "selName")
    else :
        merchantNm = getParameter(form_data , "selName")
    queryData = {
        'summSeq':'',
        'merchantId': getParameter(form_data , "submerchantId"),
        'serviceId': getParameter(form_data , "serviceId"),
        'approvalStatus': getParameter(form_data , "status"),
        'orderNo': getParameter(form_data,"orderNo"),
        'approvalNo' :getParameter(form_data,"approvalNo"),
        'dateType': getParameter(form_data,"dateType"),
        'svcConnId': getParameter(form_data,"svcConnId"),
        'startDate': startDate,
        'endDate': endDate,
        'merchantNm': merchantNm,
        'serviceNm': serviceNm,
        'offset': setStringToNumber(request.args.get("start")),
        'limit': setStringToNumber(request.args.get("length")),
        'sort':'',
        'clause':'',
        'payMethod': getParameter(form_data,"payMethod"),
        'saleDivider': getParameter(form_data,"saleDivider"),
        'cardType': getParameter(form_data,"cardType"),        
        'excelAllFlag':'',
    }
    result_data = getApiData("/approvals/chargementsMonth" ,queryData)
    return json.dumps(result_data)
    
@approvalsApi.route('/api/approvals/chargements/dailyExcelAll', methods=['GET'])
def chargementsDailyExcelAll():
    form_data = {}
    searchDate = getParameter(form_data , "startDate").split(' - ')
    selType = getParameter(form_data , "selType")
    startDate = paramEscape(searchDate[0])
    endDate = paramEscape(searchDate[1])    
    serviceNm = ""
    merchantNm = ""
    if selType == "1" :
        serviceNm = getParameter(form_data , "selName")
    else :
        merchantNm = getParameter(form_data , "selName")
        
    queryData = {
        'summSeq':'',
        'merchantId': getParameter(form_data , "submerchantId"),
        'serviceId': getParameter(form_data , "serviceId"),
        'approvalStatus': getParameter(form_data , "status"),
        'orderNo': getParameter(form_data,"orderNo"),
        'approvalNo' :getParameter(form_data,"approvalNo"),
        'dateType': getParameter(form_data,"dateType"),
        'svcConnId': getParameter(form_data,"svcConnId"),
        'startDate': startDate,
        'endDate': endDate,
        'merchantNm': merchantNm,
        'serviceNm': serviceNm,
        'offset': 0,
        'limit': EXCEL_FILE_DOWNLOAD_COUNT,
        'sort':'',
        'clause':'',
        'excelAllFlag':'1',
        'payMethod': getParameter(form_data,"payMethod"),
        'saleDivider': getParameter(form_data,"saleDivider"),
        'cardType': getParameter(form_data,"cardType"),        
        'empId' : session['empId']
    }
    rootPath = current_app.root_path
    t1 = threading.Thread(target=makeChargementsDailyExcelFile,args=[queryData,rootPath])
    t1.daemon = True
    t1.start()        
    return "엑셀 작업요청"

def makeChargementsDailyExcelFile(queryData,rootPath):
    excelZip = None
    jobStatus = 0
    batchId = None
    try :
        fileCnt = 1
        makeTime = str(int(round(time.time()*1000)))
        uploads = os.path.join(rootPath, "fileDownload" , "excel" , makeTime)
        if not os.path.isdir(uploads):
            os.makedirs(uploads)
        fileName = '거래내역_충전_일별_' +  datetime.datetime.now().strftime('%Y%m%d') + '_' + str(fileCnt) +  '.xlsx'
        zipFileName = u'거래내역_충전_일별_'+ datetime.datetime.now().strftime('%Y%m%d') +'.zip'
        #작업 시작 메시지 추가
        batchId = postBatchMng({
            "reqId"    : queryData['empId'],
            "status"   : "BAT-0001" , # 진행중
            "filePath" : os.path.join(uploads ,zipFileName),
            "content"  : "충전 일별 정산 엑셀 배치작업",
            "errMsg"   : ""
        })["data"]["batchId"]
        workbook = xlsxwriter.Workbook(os.path.join(uploads ,setUnicodeEncodeTypeToEucKr(fileName)))
        worksheet = workbook.add_worksheet()
        money_format = workbook.add_format()
        money_format.set_num_format('_- #,##0_-;[red]- #,##0_-;_- "-"_-;_-@_-')
        row = 0
        worksheet.write(row, 0  ,"거래처명")
        worksheet.write(row, 1  ,"서비스명")
        worksheet.write(row, 2  ,"연동아이디")
        worksheet.write(row, 3  ,"상태")
        worksheet.write(row, 4  ,"충전수단")
        worksheet.write(row, 5  ,"카드구분")
        worksheet.write(row, 6  ,"영업일")
        worksheet.write(row, 7  ,"거래일")
        worksheet.write(row, 8  ,"거래건수")
        worksheet.write(row, 9  ,"거래금액")
        worksheet.write(row, 10 ,"수수료")
        while True : 
            searchData = getData("/approvals/chargementsDaily" ,queryData)
            for data in searchData["resultList"]:
                row += 1
                amount = long(data["amount"])
                commision = 0
                if data["commision"] > 0 and isnumeric(data["commision"]):
                    commision = float(data["commision"])
                if data["approvalStatus"] != "CHST-0001" and data["approvalStatus"] != "CHST-0004" and data["approvalStatus"] != "CHST-0005" and data["approvalStatus"] != "CHST-0008" and data["approvalStatus"] != "CHST-0009" :
                    amount *= -1
                    if commision > 0 :
                        commision *= -1
                worksheet.write(row, 0  ,data["submerchantNm"])
                worksheet.write(row, 1  ,data["serviceNm"])
                worksheet.write(row, 2  ,data["svcConnId"])
                worksheet.write(row, 3  ,data["approvalStatusNm"])
                worksheet.write(row, 4  ,data["payTypeName"])
                worksheet.write(row, 5  ,data["cardCdName"])
                worksheet.write(row, 6  ,parseDate(data["dealDt"] ,'%Y%m%d','%Y-%m-%d'))
                worksheet.write(row, 7  ,parseDate(data["dealDt"] ,'%Y%m%d','%Y-%m-%d'))
                worksheet.write_number(row, 8, long(data["cnt"]), money_format)
                worksheet.write_number(row, 9, amount, money_format)
                worksheet.write_number(row, 10, long(commision), money_format)
                if row >= EXCEL_FILE_MAKE_LIMT_COUNT :
                    row = 0
                    fileCnt += 1
                    fileName = '거래내역_충전_일별_' +  datetime.datetime.now().strftime('%Y%m%d') + '_' + str(fileCnt) +  '.xlsx'
                    # 디비 조회건수 * 2 row 생성시 파일 재생성
                    workbook.close()
                    workbook = xlsxwriter.Workbook(os.path.join(uploads ,setUnicodeEncodeTypeToEucKr(fileName)))
                    worksheet = workbook.add_worksheet()
                    worksheet.write(row, 0  ,"거래처명")
                    worksheet.write(row, 1  ,"서비스명")
                    worksheet.write(row, 2  ,"연동아이디")
                    worksheet.write(row, 3  ,"상태")
                    worksheet.write(row, 4  ,"충전수단")
                    worksheet.write(row, 5  ,"카드구분")
                    worksheet.write(row, 6  ,"영업일")
                    worksheet.write(row, 7  ,"거래일")
                    worksheet.write(row, 8  ,"거래건수")
                    worksheet.write(row, 9  ,"거래금액")
                    worksheet.write(row, 10 ,"수수료")        
            queryData["offset"] = queryData["offset"] + EXCEL_FILE_DOWNLOAD_COUNT 
            if len(searchData["resultList"]) < EXCEL_FILE_DOWNLOAD_COUNT : 
                break
        workbook.close()
        excelZip = zipfile.ZipFile(os.path.join(uploads ,zipFileName),'w')
        for folder, subfolders, files in os.walk(uploads):
            for file in files:
                if file.endswith('.xlsx'):
                    excelZip.write(os.path.join(folder ,file), setUnicodeFormatToEucKr(file), compress_type=zipfile.ZIP_DEFLATED)
    except Exception as err:
        #에러 메시지 추가
        putBatchMng({
            "batchId"  : str(batchId),
            "reqId"    : queryData['empId'],
            "status"   : "BAT-0003" , # 진행중
            "errMsg"   : str(err)
        })
        jobStatus = 1
        print err
    finally:
        if excelZip != None:
            excelZip.close()
        #성공 메시지 추가
        if jobStatus == 0 :    
            putBatchMng({
                "batchId"  : str(batchId),
                "reqId"    : queryData['empId'],
                "status"   : "BAT-0002" , # 진행중
                "errMsg"   : ""
            })        
        print "종료"
        
@approvalsApi.route('/api/approvals/chargements/monthExcelAll', methods=['GET'])
def chargementsMonthExcelAll():
    form_data = {}
    searchDate = getParameter(form_data , "startDate").split(' - ')
    selType = getParameter(form_data , "selType")
    startDate = paramEscape(searchDate[0])
    endDate = paramEscape(searchDate[1])    
    serviceNm = ""
    merchantNm = ""
    if selType == "1" :
        serviceNm = getParameter(form_data , "selName")
    else :
        merchantNm = getParameter(form_data , "selName")
        
    queryData = {
        'summSeq':'',
        'merchantId': getParameter(form_data , "submerchantId"),
        'serviceId': getParameter(form_data , "serviceId"),
        'approvalStatus': getParameter(form_data , "status"),
        'orderNo': getParameter(form_data,"orderNo"),
        'approvalNo' :getParameter(form_data,"approvalNo"),
        'dateType': getParameter(form_data,"dateType"),
        'svcConnId': getParameter(form_data,"svcConnId"),
        'startDate': startDate,
        'endDate': endDate,
        'merchantNm': merchantNm,
        'serviceNm': serviceNm,
        'offset': 0,
        'limit': EXCEL_FILE_DOWNLOAD_COUNT,
        'sort':'',
        'clause':'',
        'excelAllFlag':'1',
        'payMethod': getParameter(form_data,"payMethod"),
        'saleDivider': getParameter(form_data,"saleDivider"),
        'cardType': getParameter(form_data,"cardType"),        
        'empId' : session['empId']
    }
    rootPath = current_app.root_path
    t1 = threading.Thread(target=makeChargementsMonthExcelFile,args=[queryData,rootPath])
    t1.daemon = True
    t1.start()        
    return "엑셀 작업요청"

def makeChargementsMonthExcelFile(queryData,rootPath):
    excelZip = None
    jobStatus = 0
    batchId = None
    try :
        fileCnt = 1
        makeTime = str(int(round(time.time()*1000)))
        uploads = os.path.join(rootPath, "fileDownload" , "excel" , makeTime)
        if not os.path.isdir(uploads):
            os.makedirs(uploads)
        fileName = '거래내역_충전_월별_' +  datetime.datetime.now().strftime('%Y%m%d') + '_' + str(fileCnt) +  '.xlsx'
        zipFileName = u'거래내역_충전_월별_'+ datetime.datetime.now().strftime('%Y%m%d') +'.zip'
        #작업 시작 메시지 추가
        batchId = postBatchMng({
            "reqId"    : queryData['empId'],
            "status"   : "BAT-0001" , # 진행중
            "filePath" : os.path.join(uploads ,zipFileName),
            "content"  : "충전 월별 정산 엑셀 배치작업",
            "errMsg"   : ""
        })["data"]["batchId"]
        workbook = xlsxwriter.Workbook(os.path.join(uploads ,setUnicodeEncodeTypeToEucKr(fileName)))
        worksheet = workbook.add_worksheet()
        money_format = workbook.add_format()
        money_format.set_num_format('_- #,##0_-;[red]- #,##0_-;_- "-"_-;_-@_-')
        row = 0
        worksheet.write(row, 0  ,"거래처명")
        worksheet.write(row, 1  ,"서비스명")
        worksheet.write(row, 2  ,"연동아이디")
        worksheet.write(row, 3  ,"상태")
        worksheet.write(row, 4  ,"충전수단")
        worksheet.write(row, 5  ,"카드구분")
        worksheet.write(row, 6  ,"영업일")
        worksheet.write(row, 7  ,"거래일")
        worksheet.write(row, 8  ,"거래건수")
        worksheet.write(row, 9  ,"거래금액")
        worksheet.write(row, 10 ,"수수료")
        while True : 
            searchData = getData("/approvals/chargementsMonth" ,queryData)
            for data in searchData["resultList"]:

                row += 1
                amount = long(data["amount"])
                commision = 0
                if data["commision"] > 0 and isnumeric(data["commision"]):
                    commision = float(data["commision"])
                if data["approvalStatus"] != "CHST-0001" and data["approvalStatus"] != "CHST-0004" and data["approvalStatus"] != "CHST-0005" and data["approvalStatus"] != "CHST-0008" and data["approvalStatus"] != "CHST-0009" :
                    amount *= -1
                    if commision > 0 :
                        commision *= -1
                worksheet.write(row, 0  ,data["submerchantNm"])
                worksheet.write(row, 1  ,data["serviceNm"])
                worksheet.write(row, 2  ,data["svcConnId"])
                worksheet.write(row, 3  ,data["approvalStatusNm"])
                worksheet.write(row, 4  ,data["payTypeName"])
                worksheet.write(row, 5  ,data["cardCdName"])
                worksheet.write(row, 6  ,parseDate(data["dealDt"] ,'%Y%m','%Y-%m'))
                worksheet.write(row, 7  ,parseDate(data["dealDt"] ,'%Y%m','%Y-%m'))
                worksheet.write_number(row, 8, long(data["cnt"]), money_format)
                worksheet.write_number(row, 9, amount, money_format)
                worksheet.write_number(row, 10, long(commision), money_format)
                if row >= EXCEL_FILE_MAKE_LIMT_COUNT :
                    row = 0
                    fileCnt += 1
                    fileName = '거래내역_충전_월별_' +  datetime.datetime.now().strftime('%Y%m%d') + '_' + str(fileCnt) +  '.xlsx'
                    # 디비 조회건수 * 2 row 생성시 파일 재생성
                    workbook.close()
                    workbook = xlsxwriter.Workbook(os.path.join(uploads ,setUnicodeEncodeTypeToEucKr(fileName)))
                    worksheet = workbook.add_worksheet()
                    worksheet.write(row, 0  ,"거래처명")
                    worksheet.write(row, 1  ,"서비스명")
                    worksheet.write(row, 2  ,"연동아이디")
                    worksheet.write(row, 3  ,"상태")
                    worksheet.write(row, 4  ,"충전수단")
                    worksheet.write(row, 5  ,"카드구분")
                    worksheet.write(row, 6  ,"영업일")
                    worksheet.write(row, 7  ,"거래일")
                    worksheet.write(row, 8  ,"거래건수")
                    worksheet.write(row, 9  ,"거래금액")
                    worksheet.write(row, 10 ,"수수료")                   
            queryData["offset"] = queryData["offset"] + EXCEL_FILE_DOWNLOAD_COUNT 
            if len(searchData["resultList"]) < EXCEL_FILE_DOWNLOAD_COUNT : 
                break
        workbook.close()
        excelZip = zipfile.ZipFile(os.path.join(uploads ,zipFileName),'w')
        for folder, subfolders, files in os.walk(uploads):
            for file in files:
                if file.endswith('.xlsx'):
                    excelZip.write(os.path.join(folder ,file), setUnicodeFormatToEucKr(file), compress_type=zipfile.ZIP_DEFLATED)
    except Exception as err:
        #에러 메시지 추가
        putBatchMng({
            "batchId"  : str(batchId),
            "reqId"    : queryData['empId'],
            "status"   : "BAT-0003" , # 진행중
            "errMsg"   : str(err)
        })
        jobStatus = 1
        print err
    finally:
        if excelZip != None:
            excelZip.close()
        #성공 메시지 추가
        if jobStatus == 0 :    
            putBatchMng({
                "batchId"  : str(batchId),
                "reqId"    : queryData['empId'],
                "status"   : "BAT-0002" , # 진행중
                "errMsg"   : ""
            })        
        print "종료"
        
        
@approvalsApi.route('/api/approvals/chargements/excelAll', methods=['GET'])
def chargementsExcelAll():
    form_data = {}
    searchDate = getParameter(form_data , "startDate").split(' - ')
    selType = getParameter(form_data , "selType")
    startDate = paramEscape(searchDate[0])
    endDate = paramEscape(searchDate[1])    
    serviceNm = ""
    merchantNm = ""
    if selType == "1" :
        serviceNm = getParameter(form_data , "selName")
    else :
        merchantNm = getParameter(form_data , "selName")
        
    queryData = {
        'summSeq':'',
        'merchantId': getParameter(form_data , "submerchantId"),
        'serviceId': getParameter(form_data , "serviceId"),
        'approvalStatus': getParameter(form_data , "status"),
        'orderNo': getParameter(form_data,"orderNo"),
        'approvalNo' :getParameter(form_data,"approvalNo"),
        'dateType': getParameter(form_data,"dateType"),
        'svcConnId': getParameter(form_data,"svcConnId"),
        'startDate': startDate,
        'endDate': endDate,
        'merchantNm': merchantNm,
        'serviceNm': serviceNm,
        'offset': 0,
        'limit': EXCEL_FILE_DOWNLOAD_COUNT,
        'sort':'',
        'clause':'',
        'excelAllFlag':'1',
        'payMethod': getParameter(form_data,"payMethod"),
        'saleDivider': getParameter(form_data,"saleDivider"),
        'cardType': getParameter(form_data,"cardType"),        
        'empId' : session['empId']
    }
    rootPath = current_app.root_path
    t1 = threading.Thread(target=makeChargementsExcelFile,args=[queryData,rootPath])
    t1.daemon = True
    t1.start()        
    return "엑셀 작업요청"

def makeChargementsExcelFile(queryData,rootPath):
    excelZip = None
    jobStatus = 0
    batchId = None
    try :
        fileCnt = 1
        makeTime = str(int(round(time.time()*1000)))
        uploads = os.path.join(rootPath, "fileDownload" , "excel" , makeTime)
        if not os.path.isdir(uploads):
            os.makedirs(uploads)
        fileName = '거래내역_충전_' +  datetime.datetime.now().strftime('%Y%m%d') + '_' + str(fileCnt) +  '.xlsx'
        zipFileName = u'거래내역_충전_'+ datetime.datetime.now().strftime('%Y%m%d') +'.zip'
        #작업 시작 메시지 추가
        batchId = postBatchMng({
            "reqId"    : queryData['empId'],
            "status"   : "BAT-0001" , # 진행중
            "filePath" : os.path.join(uploads ,zipFileName),
            "content"  : "충전 정산 엑셀 배치작업",
            "errMsg"   : ""
        })["data"]["batchId"]
        workbook = xlsxwriter.Workbook(os.path.join(uploads ,setUnicodeEncodeTypeToEucKr(fileName)))
        worksheet = workbook.add_worksheet()
        money_format = workbook.add_format()
        money_format.set_num_format('_- #,##0_-;[red]- #,##0_-;_- "-"_-;_-@_-')
        row = 0
        worksheet.write(row, 0  ,"거래처명")
        worksheet.write(row, 1  ,"서비스명")
        worksheet.write(row, 2  ,"연동아이디")
        worksheet.write(row, 3  ,"점포코드")
        worksheet.write(row, 4  ,"점포명(사용처)")
        worksheet.write(row, 5  ,"상태")
        worksheet.write(row, 6  ,"구분")
        worksheet.write(row, 7  ,"충전수단")
        worksheet.write(row, 8  ,"카드구분")
        worksheet.write(row, 9  ,"영업일")
        worksheet.write(row, 10  ,"거래일")
        worksheet.write(row, 11 ,"거래시간")
        worksheet.write(row, 12 ,"주문번호")
        worksheet.write(row, 13 ,"승인번호")
        worksheet.write(row, 14 ,"카드번호")
        worksheet.write(row, 15 ,"거래금액")
        worksheet.write(row, 16 ,"수수료")
        worksheet.write(row, 17 ,"비고")
        while True : 
            searchData = getData("/approvals/chargements" ,queryData)
            for data in searchData["resultList"]:
                row += 1
                amount = long(data["amount"])
                commision = 0
                if data["commision"] > 0 and isnumeric(data["commision"]):
                    commision = float(data["commision"])
                if data["approvalStatus"] != "CHST-0001" and data["approvalStatus"] != "CHST-0004" and data["approvalStatus"] != "CHST-0005" and data["approvalStatus"] != "CHST-0008" and data["approvalStatus"] != "CHST-0009" :
                    amount *= -1
                    if commision > 0 :
                        commision *= -1
                          
                worksheet.write(row, 0  ,data["submerchantNm"])
                worksheet.write(row, 1  ,data["serviceNm"])
                worksheet.write(row, 2  ,data["svcConnId"])
                worksheet.write(row, 3  ,data["storeCd"])
                worksheet.write(row, 4  ,data["storeNm"])
                worksheet.write(row, 5  ,data["approvalStatusNm"])
                worksheet.write(row, 6  ,data["chargeStatusName"])
                worksheet.write(row, 7  ,data["chargeMethodName"])
                worksheet.write(row, 8  ,data["cardCdName"])
                worksheet.write(row, 9  ,parseDate(data["dealDt"] ,'%Y%m%d','%Y-%m-%d'))
                worksheet.write(row, 10 ,parseDate(data["approvalDt"] ,'%Y-%m-%d %H:%M:%S' ,'%Y-%m-%d'))
                worksheet.write(row, 11 ,parseDate(data["approvalDt"] ,'%Y-%m-%d %H:%M:%S' ,'%H:%M:%S'))
                worksheet.write(row, 12 ,data["orderNo"])
                worksheet.write(row, 13 ,data["approvalNo"])
                worksheet.write(row, 14 ,data["cardNo"])
                worksheet.write_number(row, 15, amount, money_format)
                worksheet.write_number(row, 16, long(commision), money_format)
                worksheet.write(row, 17 ,data["desc01"])
                if row >= EXCEL_FILE_MAKE_LIMT_COUNT :
                    row = 0
                    fileCnt += 1
                    fileName = '거래내역_충전_' +  datetime.datetime.now().strftime('%Y%m%d') + '_' + str(fileCnt) +  '.xlsx'
                    # 디비 조회건수 * 2 row 생성시 파일 재생성
                    workbook.close()
                    workbook = xlsxwriter.Workbook(os.path.join(uploads ,setUnicodeEncodeTypeToEucKr(fileName)))
                    worksheet = workbook.add_worksheet()
                    worksheet.write(row, 0  ,"거래처명")
                    worksheet.write(row, 1  ,"서비스명")
                    worksheet.write(row, 2  ,"연동아이디")
                    worksheet.write(row, 3  ,"점포코드")
                    worksheet.write(row, 4  ,"점포명(사용처)")
                    worksheet.write(row, 5  ,"상태")
                    worksheet.write(row, 6  ,"구분")
                    worksheet.write(row, 7  ,"충전수단")
                    worksheet.write(row, 8  ,"카드구분")
                    worksheet.write(row, 9  ,"영업일")
                    worksheet.write(row, 10  ,"거래일")
                    worksheet.write(row, 11 ,"거래시간")
                    worksheet.write(row, 12 ,"주문번호")
                    worksheet.write(row, 13 ,"승인번호")
                    worksheet.write(row, 14 ,"카드번호")
                    worksheet.write(row, 15 ,"거래금액")
                    worksheet.write(row, 16 ,"수수료")                               
                    worksheet.write(row, 17 ,"비고")                               
            queryData["offset"] = queryData["offset"] + EXCEL_FILE_DOWNLOAD_COUNT 
            if len(searchData["resultList"]) < EXCEL_FILE_DOWNLOAD_COUNT : 
                break
        workbook.close()
        excelZip = zipfile.ZipFile(os.path.join(uploads ,zipFileName),'w')
        for folder, subfolders, files in os.walk(uploads):
            for file in files:
                if file.endswith('.xlsx'):
                    excelZip.write(os.path.join(folder ,file), setUnicodeFormatToEucKr(file), compress_type=zipfile.ZIP_DEFLATED)
    except Exception as err:
        #에러 메시지 추가
        putBatchMng({
            "batchId"  : str(batchId),
            "reqId"    : queryData['empId'],
            "status"   : "BAT-0003" , # 진행중
            "errMsg"   : str(err)
        })
        jobStatus = 1
        print err
    finally:
        if excelZip != None:
            excelZip.close()
        #성공 메시지 추가
        if jobStatus == 0 :    
            putBatchMng({
                "batchId"  : str(batchId),
                "reqId"    : queryData['empId'],
                "status"   : "BAT-0002" , # 진행중
                "errMsg"   : ""
            })        
        print "종료"        

@approvalsApi.route('/api/approvals/salements', methods=['GET'])
def salements():
    form_data = json.loads(request.args.get("formData"))
    searchDate = getParameter(form_data , "startDate").split(' - ')
    selType = getParameter(form_data , "selType")
    noType = getParameter(form_data , "noType")
    startDate = paramEscape(searchDate[0])
    endDate = paramEscape(searchDate[1])    
    queryData = {
        'merchantId': getParameter(form_data , "submerchantId"),
        'serviceId': getParameter(form_data , "serviceId"),
        'approvalStatus': getParameter(form_data , "status"),
        'saleMethod': getParameter(form_data,"saleTypeDetail"),
        'prodType': getParameter(form_data,"saleType"),
        'orderNo': noType == "1" and getParameter(form_data, "approvalNo") or "",
        'approvalNo' : noType == "2" and getParameter(form_data, "approvalNo") or "",
        'svcConnId': getParameter(form_data,"svcConnId"),
        'amount': paramEscape(getParameter(form_data,"amount")),
        'startDate': startDate,
        'endDate': endDate,
        'merchantNm': selType == "0" and getParameter(form_data, "selName") or "",
        'serviceNm': selType == "1" and getParameter(form_data, "selName") or "",
        'offset': setStringToNumber(request.args.get("start")),
        'limit': setStringToNumber(request.args.get("length")),
        'excelAllFlag':'',
    }
    result_data = getApiData("/approvals/salements" ,queryData)
    return json.dumps(result_data)

@approvalsApi.route('/api/approvals/salementsMonth', methods=['GET'])
def salementsMonth():
    form_data = json.loads(request.args.get("formData"))
    searchDate = getParameter(form_data , "startDate").split(' - ')
    selType = getParameter(form_data , "selType")
    noType = getParameter(form_data , "noType")
    startDate = paramEscape(searchDate[0])
    endDate = paramEscape(searchDate[1])    
    queryData = {
        'merchantId': getParameter(form_data , "submerchantId"),
        'serviceId': getParameter(form_data , "serviceId"),
        'approvalStatus': getParameter(form_data , "status"),
        'saleMethod': getParameter(form_data,"saleTypeDetail"),
        'prodType': getParameter(form_data,"saleType"),
        'orderNo': noType == "1" and getParameter(form_data, "approvalNo") or "",
        'approvalNo' : noType == "2" and getParameter(form_data, "approvalNo") or "",
        'svcConnId': getParameter(form_data,"svcConnId"),
        'amount': paramEscape(getParameter(form_data,"amount")),
        'startDate': startDate,
        'endDate': endDate,
        'merchantNm': selType == "0" and getParameter(form_data, "selName") or "",
        'serviceNm': selType == "1" and getParameter(form_data, "selName") or "",
        'offset': setStringToNumber(request.args.get("start")),
        'limit': setStringToNumber(request.args.get("length")),
        'excelAllFlag':'',
    }
    result_data = getApiData("/approvals/salementsMonth" ,queryData)
    return json.dumps(result_data)

@approvalsApi.route('/api/approvals/salements/excelAll', methods=['GET'])
def salementsExcelAll():
    form_data = {}
    searchDate = getParameter(form_data , "startDate").split(' - ')
    selType = getParameter(form_data , "selType")
    noType = getParameter(form_data , "noType")
    startDate = paramEscape(searchDate[0])
    endDate = paramEscape(searchDate[1])    
    queryData = {
        'merchantId': getParameter(form_data , "submerchantId"),
        'serviceId': getParameter(form_data , "serviceId"),
        'approvalStatus': getParameter(form_data , "status"),
        'saleMethod': getParameter(form_data,"saleTypeDetail"),
        'prodType': getParameter(form_data,"saleType"),
        'orderNo': noType == "1" and getParameter(form_data, "approvalNo") or "",
        'approvalNo' : noType == "2" and getParameter(form_data, "approvalNo") or "",
        'svcConnId': getParameter(form_data,"svcConnId"),
        'amount': paramEscape(getParameter(form_data,"amount")),
        'startDate': startDate,
        'endDate': endDate,
        'merchantNm': selType == "0" and getParameter(form_data, "selName") or "",
        'serviceNm': selType == "1" and getParameter(form_data, "selName") or "",
        'offset': 0,
        'limit': EXCEL_FILE_DOWNLOAD_COUNT,
        'excelAllFlag': '1',
        'empId' : session['empId']
    }

    rootPath = current_app.root_path
    t1 = threading.Thread(target=makeSalementsExcelFile,args=[queryData,rootPath])
    t1.daemon = True
    t1.start()
    
    return "엑셀 작업요청"

def makeSalementsExcelFile(queryData,rootPath):
    excelZip = None
    jobStatus = 0
    batchId = None
    try:
        fileCnt = 1
        makeTime = str(int(round(time.time()*1000)))
        uploads = os.path.join(rootPath, "fileDownload" , "excel" , makeTime)
        if not os.path.isdir(uploads):
            os.makedirs(uploads)
        zipFileName = u'판매정산내역_'+ datetime.datetime.now().strftime('%Y%m%d') +'.zip'
        #작업 시작 메시지 추가
        batchId = postBatchMng({
            "reqId"    : queryData['empId'],
            "status"   : "BAT-0001" , # 진행중
            "filePath" : os.path.join(uploads ,zipFileName),
            "content"  : "판매 정산 엑셀 배치작업",
            "errMsg"   : ""
        })["data"]["batchId"]            
        fileName = '판매정산내역_' + datetime.datetime.now().strftime('%Y%m%d') + '_' + str(fileCnt) +  '.xlsx'
        workbook = xlsxwriter.Workbook(os.path.join(uploads ,setUnicodeEncodeTypeToEucKr(fileName)))
        worksheet = workbook.add_worksheet()
        money_format = workbook.add_format()
        money_format.set_num_format('_- #,##0_-;[red]- #,##0_-;_- "-"_-;_-@_-')
        row = 0
        worksheet.write(row, 0  ,"거래처명")
        worksheet.write(row, 1  ,"서비스명")
        worksheet.write(row, 2  ,"연동아이디")
        worksheet.write(row, 3  ,"상태")
        worksheet.write(row, 4  ,"상품명")
        worksheet.write(row, 5  ,"상품ID")
        worksheet.write(row, 6  ,"상품구분")
        worksheet.write(row, 7  ,"거래일")
        worksheet.write(row, 8  ,"거래시간")
        worksheet.write(row, 9  ,"주문번호")
        worksheet.write(row, 10 ,"승인번호")
        worksheet.write(row, 11 ,"권종")
        worksheet.write(row, 12 ,"거래건수")
        worksheet.write(row, 13 ,"거래금액")
        worksheet.write(row, 14 ,"비고")
        while True : 
            searchData = getData("/approvals/salements" ,queryData)
            for data in searchData["resultList"]:
                row += 1
                saleCnt = long(data["saleCnt"])
                saleAmt = long(data["saleAmt"]) * saleCnt
                if data["approvalStatus"] == "SSTS-0002":
                    saleAmt = saleAmt * -1
                worksheet.write(row, 0  ,data["submerchantNm"])
                worksheet.write(row, 1  ,data["serviceNm"])
                worksheet.write(row, 2  ,data["svcConnId"])
                worksheet.write(row, 3  ,data["approvalStatusNm"])
                worksheet.write(row, 4  ,data["prodNm"])
                worksheet.write(row, 5  ,data["prodCd"])
                worksheet.write(row, 6  ,data["prodTypeNm"])
                worksheet.write(row, 7  ,parseDate(data["approvalDt"] ,'%Y-%m-%d %H:%M:%S' ,'%Y-%m-%d'))
                worksheet.write(row, 8  ,parseDate(data["approvalDt"] ,'%Y-%m-%d %H:%M:%S' ,'%H:%M:%S'))
                worksheet.write(row, 9  ,data["orderNo"])
                worksheet.write(row, 10 ,data["approvalNo"])
                worksheet.write_number(row, 11, long(data["saleAmt"]), money_format)
                worksheet.write_number(row, 12, saleCnt, money_format)
                worksheet.write_number(row, 13, saleAmt, money_format)
                worksheet.write(row, 14 ,data["desc01"])
                if row >= EXCEL_FILE_MAKE_LIMT_COUNT :
                    row = 0
                    fileCnt += 1
                    fileName = '판매정산내역_' + datetime.datetime.now().strftime('%Y%m%d') + '_' + str(fileCnt) +  '.xlsx'
                    # 디비 조회건수 * 2 row 생성시 파일 재생성
                    workbook.close()
                    workbook = xlsxwriter.Workbook(os.path.join(uploads ,setUnicodeEncodeTypeToEucKr(fileName)))
                    worksheet = workbook.add_worksheet()
                    worksheet.write(row, 0  ,"거래처명")
                    worksheet.write(row, 1  ,"서비스명")
                    worksheet.write(row, 2  ,"연동아이디")
                    worksheet.write(row, 3  ,"상태")
                    worksheet.write(row, 4  ,"상품명")
                    worksheet.write(row, 5  ,"상품ID")
                    worksheet.write(row, 6  ,"상품구분")
                    worksheet.write(row, 7  ,"거래일")
                    worksheet.write(row, 8  ,"거래시간")
                    worksheet.write(row, 9  ,"주문번호")
                    worksheet.write(row, 10 ,"승인번호")
                    worksheet.write(row, 11 ,"권종")
                    worksheet.write(row, 12 ,"거래건수")
                    worksheet.write(row, 13 ,"거래금액")
                    worksheet.write(row, 14 ,"비고")                              
            queryData["offset"] = queryData["offset"] + EXCEL_FILE_DOWNLOAD_COUNT 
            if len(searchData["resultList"]) < EXCEL_FILE_DOWNLOAD_COUNT : 
                break
        workbook.close()
        excelZip = zipfile.ZipFile(os.path.join(uploads ,zipFileName),'w')
        for folder, subfolders, files in os.walk(uploads):
            for file in files:
                if file.endswith('.xlsx'):
                    excelZip.write(os.path.join(folder ,file), setUnicodeFormatToEucKr(file), compress_type=zipfile.ZIP_DEFLATED)
                    
    except Exception as err:
        #에러 메시지 추가
        putBatchMng({
            "batchId"  : str(batchId),
            "reqId"    : queryData['empId'],
            "status"   : "BAT-0003" , # 진행중
            "errMsg"   : str(err)
        })
        jobStatus = 1
        print err
    finally:
        if excelZip != None:
            excelZip.close()
        if jobStatus == 0 :    
            putBatchMng({
                "batchId"  : str(batchId),
                "reqId"    : queryData['empId'],
                "status"   : "BAT-0002" , # 진행중
                "errMsg"   : ""
            })               
        #성공 메시지 추가
        print "성공"
        
@approvalsApi.route('/api/approvals/salements/monthExcelAll', methods=['GET'])
def salementsMonthExcelAll():
    form_data = {}
    searchDate = getParameter(form_data , "startDate").split(' - ')
    selType = getParameter(form_data , "selType")
    noType = getParameter(form_data , "noType")
    startDate = paramEscape(searchDate[0])
    endDate = paramEscape(searchDate[1])    
    queryData = {
        'merchantId': getParameter(form_data , "submerchantId"),
        'serviceId': getParameter(form_data , "serviceId"),
        'approvalStatus': getParameter(form_data , "status"),
        'saleMethod': getParameter(form_data,"saleTypeDetail"),
        'prodType': getParameter(form_data,"saleType"),
        'orderNo': noType == "1" and getParameter(form_data, "approvalNo") or "",
        'approvalNo' : noType == "2" and getParameter(form_data, "approvalNo") or "",
        'svcConnId': getParameter(form_data,"svcConnId"),
        'amount': paramEscape(getParameter(form_data,"amount")),
        'startDate': startDate,
        'endDate': endDate,
        'merchantNm': selType == "0" and getParameter(form_data, "selName") or "",
        'serviceNm': selType == "1" and getParameter(form_data, "selName") or "",
        'offset': 0,
        'limit': EXCEL_FILE_DOWNLOAD_COUNT,
        'excelAllFlag': '1',
        'empId' : session['empId']
    }

    rootPath = current_app.root_path
    t1 = threading.Thread(target=makeSalementsMonthExcelFile,args=[queryData,rootPath])
    t1.daemon = True
    t1.start()
    
    return "엑셀 작업요청"

def makeSalementsMonthExcelFile(queryData,rootPath):
    excelZip = None
    jobStatus = 0
    batchId = None
    try:
        fileCnt = 1
        makeTime = str(int(round(time.time()*1000)))
        uploads = os.path.join(rootPath, "fileDownload" , "excel" , makeTime)
        if not os.path.isdir(uploads):
            os.makedirs(uploads)
        zipFileName = u'판매정산내역_월별_'+ datetime.datetime.now().strftime('%Y%m%d') +'.zip'
        #작업 시작 메시지 추가
        batchId = postBatchMng({
            "reqId"    : queryData['empId'],
            "status"   : "BAT-0001" , # 진행중
            "filePath" : os.path.join(uploads ,zipFileName),
            "content"  : "판매 정산 월별 요약 엑셀 배치작업",
            "errMsg"   : ""
        })["data"]["batchId"]            
        fileName = '판매정산내역_월별_' +  datetime.datetime.now().strftime('%Y%m%d') + '_' + str(fileCnt) +  '.xlsx'
        workbook = xlsxwriter.Workbook(os.path.join(uploads ,setUnicodeEncodeTypeToEucKr(fileName)))
        worksheet = workbook.add_worksheet()
        money_format = workbook.add_format()
        money_format.set_num_format('_- #,##0_-;[red]- #,##0_-;_- "-"_-;_-@_-')
        row = 0
        worksheet.write(row, 0  ,"거래처명")
        worksheet.write(row, 1  ,"서비스명")
        worksheet.write(row, 2  ,"연동아이디")
        worksheet.write(row, 3  ,"상태")
        worksheet.write(row, 4  ,"상품명")
        worksheet.write(row, 5  ,"상품ID")
        worksheet.write(row, 6  ,"상품구분")
        worksheet.write(row, 7  ,"권종")
        worksheet.write(row, 8  ,"거래일")
        worksheet.write(row, 9  ,"거래건수")
        worksheet.write(row, 10 ,"거래금액")
        while True : 
            searchData = getData("/approvals/salementsMonth" ,queryData)
            for data in searchData["resultList"]:
                row += 1
                saleCnt = long(data["saleCnt"])
                saleAmt = long(data["saleAmt"]) * saleCnt
                if data["approvalStatus"] == "SSTS-0002":
                    saleAmt = saleAmt * -1
                worksheet.write(row, 0  ,data["submerchantNm"])
                worksheet.write(row, 1  ,data["serviceNm"])
                worksheet.write(row, 2  ,data["svcConnId"])
                worksheet.write(row, 3  ,data["approvalStatusNm"])
                worksheet.write(row, 4  ,data["prodNm"])
                worksheet.write(row, 5  ,data["prodCd"])
                worksheet.write(row, 6  ,data["prodTypeNm"])
                worksheet.write_number(row, 7, long(data["saleAmt"]), money_format)
                worksheet.write(row, 8  ,parseDate(data["dealDt"] + "01" ,'%Y%m%d' ,'%Y-%m-%d'))
                worksheet.write_number(row, 9, saleCnt, money_format)
                worksheet.write_number(row, 10, saleAmt, money_format)
                if row >= EXCEL_FILE_MAKE_LIMT_COUNT :
                    row = 0
                    fileCnt += 1
                    fileName = '판매정산내역_월별_' +  datetime.datetime.now().strftime('%Y%m%d') + '_' + str(fileCnt) +  '.xlsx'
                    # 디비 조회건수 * 2 row 생성시 파일 재생성
                    workbook.close()
                    workbook = xlsxwriter.Workbook(os.path.join(uploads ,setUnicodeEncodeTypeToEucKr(fileName)))
                    worksheet = workbook.add_worksheet()
                    worksheet.write(row, 0  ,"거래처명")
                    worksheet.write(row, 1  ,"서비스명")
                    worksheet.write(row, 2  ,"연동아이디")
                    worksheet.write(row, 3  ,"상태")
                    worksheet.write(row, 4  ,"상품명")
                    worksheet.write(row, 5  ,"상품ID")
                    worksheet.write(row, 6  ,"상품구분")
                    worksheet.write(row, 7  ,"권종")
                    worksheet.write(row, 8  ,"거래일")
                    worksheet.write(row, 9  ,"거래건수")
                    worksheet.write(row, 10 ,"거래금액")                            
            queryData["offset"] = queryData["offset"] + EXCEL_FILE_DOWNLOAD_COUNT 
            if len(searchData["resultList"]) < EXCEL_FILE_DOWNLOAD_COUNT : 
                break
        workbook.close()
        excelZip = zipfile.ZipFile(os.path.join(uploads ,zipFileName),'w')
        for folder, subfolders, files in os.walk(uploads):
            for file in files:
                if file.endswith('.xlsx'):
                    excelZip.write(os.path.join(folder ,file), setUnicodeFormatToEucKr(file), compress_type=zipfile.ZIP_DEFLATED)
                    
    except Exception as err:
        #에러 메시지 추가
        putBatchMng({
            "batchId"  : str(batchId),
            "reqId"    : queryData['empId'],
            "status"   : "BAT-0003" , # 진행중
            "errMsg"   : str(err)
        })
        jobStatus = 1
        print err
    finally:
        if excelZip != None:
            excelZip.close()
        if jobStatus == 0 :    
            putBatchMng({
                "batchId"  : str(batchId),
                "reqId"    : queryData['empId'],
                "status"   : "BAT-0002" , # 진행중
                "errMsg"   : ""
            })               
        #성공 메시지 추가
        print "성공"        