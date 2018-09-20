# -- coding:utf-8 --
'''
Created on 2017. 8. 03.

@author: sanghyun
'''
import datetime
import os
import sys
import threading
import time

from flask import Blueprint, request
from flask.globals import current_app, session
import openpyxl

from routes.api.systemMngApi import postBatchMng, putBatchMng
from util.common import getData, API_SERVER_KPC_PAYMENT, paramEscape


salesApi = Blueprint('salesApi', __name__)

@salesApi.route('/api/sales/dailySales', methods=['GET'])
def test():
    queryData = {
        'searchDate': paramEscape(request.args.get("searchDate")),
        'empId' : session['empId']
    }
    print queryData
    rootPath = current_app.root_path
    t1 = threading.Thread(target=makeAggregateExcelFile,args=[queryData,rootPath])
    t1.daemon = True
    t1.start()
    
    return "엑셀 작업요청"

def makeAggregateExcelFile(queryData,rootPath):
    jobStatus = 0
    batchId = None
    try :
        uploads = os.path.join(rootPath, "docs" , "excel"  , "aggregate")
        if not os.path.isdir(uploads):
            os.makedirs(uploads)
        makeTime = str(int(round(time.time()*1000)))
        outputPath = os.path.join(rootPath, "fileDownload" , "excel" , makeTime)
        if not os.path.isdir(outputPath):
            os.makedirs(outputPath)        
        fileName = u'aggregate.xlsx'
        saveFileName = u'매출데이터' + datetime.datetime.now().strftime('%Y%m%d') + '.xlsx'
        #작업 시작 메시지 추가
        batchId = postBatchMng({
            "reqId"    : queryData['empId'],
            "status"   : "BAT-0001" , # 진행중
            "filePath" : os.path.join(outputPath ,saveFileName),
            "content"  : "매출데이터 생성",
            "errMsg"   : ""
        })["data"]["batchId"]
        searchData = getData("/api/admin/v1/aggreagate" ,queryData , API_SERVER_KPC_PAYMENT)
        excel_document = openpyxl.load_workbook(filename=os.path.join(uploads ,fileName))
        sheet = excel_document.get_sheet_by_name('sales')
        row = 2;
#         dailySum = {}
        sheet.cell(row=1 , column=7 , value =  queryData["searchDate"]  )
        # 매출 집계
        if searchData["aggregateDaily"] != None :
            for data in searchData["aggregateDaily"]:
                sheet.cell(row=row , column=1 , value = data["gubun"]      )
                sheet.cell(row=row , column=2 , value = data["name"]       )
                sheet.cell(row=row , column=3 , value = data["aplonLineId"])
                sheet.cell(row=row , column=4 , value = data["payMethod"]  )
                sheet.cell(row=row , column=5 , value = data["barCode"]    )
                sheet.cell(row=row , column=6 , value = data["payKind"]    )
                sheet.cell(row=row , column=7 , value = data["totalAmount"]  )
                row = row +1
#                 if data["aplonLineId"] in dailySum :
#                     dailySum[data["aplonLineId"]] = dailySum[data["aplonLineId"]] +  data["payAmount"]
#                 else : 
#                     dailySum[data["aplonLineId"]] = data["payAmount"]
        # 해피캐시 
        if searchData["aggregateDailyHappyCash"] != None:
            sheet.cell(row=row, column=1 , value=searchData["aggregateDailyHappyCash"]["gubun"]      )
            sheet.cell(row=row, column=2 , value="해피캐시 전환")
            sheet.cell(row=row, column=4 , value=searchData["aggregateDailyHappyCash"]["aplonLineId"])
            sheet.cell(row=row, column=3 , value=searchData["aggregateDailyHappyCash"]["payMethod"]  )
            sheet.cell(row=row, column=5 , value=searchData["aggregateDailyHappyCash"]["barCode"]    )
            sheet.cell(row=row, column=6 , value=searchData["aggregateDailyHappyCash"]["payKind"]    )
            sheet.cell(row=row, column=7 , value=searchData["aggregateDailyHappyCash"]["totalAmount"]  )
            # 우측 해피 캐시 
#             sheet.cell(row = 6 , column=11 , value=searchData["aggregateDailyHappyCash"]["payAmount"])  
            
        # 우측 데이터 GS25 POP 
#         if "gspopcards" in dailySum  : # GS25 pop 충전  
#             sheet.cell(row = 3 , column=11 , value=dailySum["gspopcards"])  
#         if "gspopcardp" in dailySum : # GS25 결제
#             sheet.cell(row = 4 , column=11 , value=dailySum["gspopcardp"])
#         gs25Etc = 0
#         if "gspopcardb2b" in dailySum : 
#             gs25Etc = gs25Etc + dailySum["gspopcardb2b"]
#         if "gswork" in dailySum : 
#             gs25Etc = gs25Etc + dailySum["gswork"]  
#         if "dmlife" in dailySum : 
#             gs25Etc = gs25Etc + dailySum["dmlife"]  
#         sheet.cell(row = 7 , column=11 , value=gs25Etc)  
#         # 우측 데이터 GSSM POP
#         if "gssmpops" in dailySum : # GSSM 충전  
#             sheet.cell(row = 8 , column=11 , value=dailySum["gssmpops"])  
#         if "gssmpopp" in dailySum : # GSSM 결제
#             sheet.cell(row = 9 , column=11 , value=dailySum["gssmpopp"])
#         # Mobile POP (*App 내 충전)  
#         if "srpoprbank" in dailySum : # M POP_계좌
#             sheet.cell(row = 11 , column=11 , value=dailySum["srpoprbank"])
#         if "srpopcard" in dailySum : # M POP_카드 
#             sheet.cell(row = 12 , column=11 , value=dailySum["srpopcard"])
#         if "kcppayco" in dailySum : # 간편결제_신용 
#             sheet.cell(row = 13 , column=11 , value=dailySum["kcppayco"])
#         if "checkpay" in dailySum : # 간편결제_계좌 
#             sheet.cell(row = 14 , column=11 , value=dailySum["checkpay"])
#         if "gsrpoppoint" in dailySum : # M POP_포인트 
#             sheet.cell(row = 15 , column=11 , value=dailySum["gsrpoppoint"])
#         if "hmcharge" in dailySum : # 상품권 충전 
#             sheet.cell(row = 16 , column=11 , value=dailySum["hmcharge"])
#         if "popcharge" in dailySum : # POP 충전권 
#             sheet.cell(row = 17 , column=11 , value=dailySum["popcharge"])
#         if "poppluscharge" in dailySum : # POP 재충전권 
#             sheet.cell(row = 18 , column=11 , value=dailySum["poppluscharge"])
#         if "happymoney_pop" in dailySum : # M_온라인 결제 
#             sheet.cell(row = 19 , column=11 , value=dailySum["happymoney_pop"])
#         # 우측 데이터 기타 충전
#         if "nasmedia" in dailySum : # 무료충전소(나스미디어)
#             sheet.cell(row = 21 , column=11 , value=dailySum["nasmedia"])
#         if "nasmedia" in dailySum : # 무료충전소(나스미디어)
#             sheet.cell(row = 22 , column=11 , value=dailySum["nasmedia"])
#         if "gsmobilepop" in dailySum : # M POP_이벤트 캐시백
#             sheet.cell(row = 23 , column=11 , value=dailySum["gsmobilepop"])
#         if "popcenter" in dailySum : # popcenter_고객센터
#             sheet.cell(row = 24 , column=11 , value=dailySum["popcenter"])
#         if "happymall" in dailySum : # 해피전산(Test)
#             sheet.cell(row = 25 , column=11 , value=dailySum["happymall"])
        
#         # 매출 집계 잔액  환불 & poppay
#         if searchData["aggregateDailyEtc"] != None :
#             for data in searchData["aggregateDailyEtc"]: 
#                 aplonLineId = data["aplonLineId"]
#                 if aplonLineId == "gspopcardp" :
#                     sheet.cell(row = 5 , column=11 , value=data["payAmount"]) # GS25 잔액환불 
#                 if aplonLineId == "gssmpopp" :
#                     sheet.cell(row = 10 , column=11 , value=data["payAmount"]) # GSSM 잔액환불
#                 if aplonLineId == "poppay" :
#                     sheet.cell(row = 26 , column=11 , value=data["payAmount"]) # poppay
        # sheet 2 요일별 데이터 
        # 모바일팝 제휴처 일별 정산대상
        sheet2 = excel_document.get_sheet_by_name('monthly')
        row = 4
        if searchData["aggregateMerchantBilling"] != None :
            for data in searchData["aggregateMerchantBilling"]:
                sheet2.cell(row=row, column=2 , value=data["tMemo"]        )
                sheet2.cell(row=row, column=3 , value=data["aplDate"]      )
                sheet2.cell(row=row, column=4 , value=data["cpId"]         )
                sheet2.cell(row=row, column=5 , value=data["svcConnId"]    )
                sheet2.cell(row=row, column=6 , value=data["paymentCount"] )
                sheet2.cell(row=row, column=7 , value=data["paymentAmount"])
                sheet2.cell(row=row, column=8 , value=data["cancelCount"]  )
                sheet2.cell(row=row, column=9 , value=data["cancelAmount"] )
                row += 1
             
        # 체크페이 
        row = 4
        if searchData["aggregateCheckPay"] != None :
            for data in searchData["aggregateCheckPay"]:
                sheet2.cell(row=row, column=14 , value=data["userId"]       )
                sheet2.cell(row=row, column=15 , value=data["aplDate"]      )
                sheet2.cell(row=row, column=16 , value=data["quantity"]     )
                sheet2.cell(row=row, column=17 , value=data["amount"]       )
                sheet2.cell(row=row, column=18 , value=data["userGubun"]    )
                sheet2.cell(row=row, column=19 , value=data["chargeType"]   )
                row = row +1
        # aggregateMonthGsPopCard
        row = 4
        if searchData["aggregateMonthGsPopCard"] != None :
            for data in searchData["aggregateMonthGsPopCard"]:
                sheet2.cell(row=row, column=23 , value=data["salesDate"]  )
                sheet2.cell(row=row, column=26 , value=data["totCount"]   )
                sheet2.cell(row=row, column=27 , value=data["totAmt"]     )
                row = row +1
        # aggregateMonthGsSmPop
        row = 37
        if searchData["aggregateMonthGsSmPop"] != None :
            for data in searchData["aggregateMonthGsSmPop"]:
                sheet2.cell(row=row, column=23 , value=data["salesDate"]  )
                sheet2.cell(row=row, column=26 , value=data["totCount"]   )
                sheet2.cell(row=row, column=27 , value=data["totAmt"]     )  
                row = row +1
                
        # aggregateMonthGshb
        row = 70
        print(searchData.get("aggregateMonthGshb"))
        if searchData["aggregateMonthGshb"] != None :
            for data in searchData["aggregateMonthGshb"]:
                sheet2.cell(row=row, column=23 , value=data["salesDate"]  )
                sheet2.cell(row=row, column=26 , value=data["totCount"]   )
                sheet2.cell(row=row, column=27 , value=data["totAmt"]     )  
                row = row +1                
        excel_document.save(filename=os.path.join(outputPath , saveFileName))
    except Exception as err:
        #에러 메시지 추가
        putBatchMng({
            "batchId"  : str(batchId),
            "reqId"    : queryData['empId'],
            "status"   : "BAT-0003" , # 진행중
            "errMsg"   : str(err)
        })
        jobStatus = 1
        print err
        print sys.exc_info()[-1].tb_lineno
    finally:
        #성공 메시지 추가
        if jobStatus == 0 :    
            putBatchMng({
                "batchId"  : str(batchId),
                "reqId"    : queryData['empId'],
                "status"   : "BAT-0002" , # 진행중
                "errMsg"   : ""
            })        
        print "종료"